from bs4 import BeautifulSoup
import re
import requests
from playwright.sync_api import sync_playwright
import lxml
import pandas as pd
import random
import time
from urllib.parse import urlparse
import xml.etree.ElementTree as Et
from openpyxl import load_workbook

# List of User-Agents. I suggest to use desktop User-Agents .
USER_AGENTS = [
    # Desktop User-Agents.
]

email_regex = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}"
phone_regex = r"(\+?\d{1,3})?(?:[ \.-]?)?(((\(?\d{4}\)?|\d{4})(?:[ \.-]?)?((\d{3})(?:[ \.-]?)?(\d{3})))|((\(?\d{3}\)?|\d{3})(?:[ \.-]?)?(((\d{4})(?:[ \.-]?)?(\d{3}))|(\d{3})(?:[ \.-]?)?(\d{4}|(\d{2}( |-|\.)?\d{2})))))"  # Updated


# Old phone regex. I'm keeping these just in case but you can delete them.
# phone_regex = r"(\+?( |-|\.)?\d{1,2}( |-|\.)?)?(\(?\d{3}\)?|\d{3})( |-|\.)?(\d{3}( |-|\.)?\d{4})"
# phone_regex = r"(\+?\d{1,3})?(?:[ \.-]?)?(\(?\d{3}\)?|\d{3})(?:[ \.-]?)?(((\d{4})(?:[ \.-]?)?(\d{3}))|(\d{3})(?:[ \.-]?)?(\d{4}|(\d{2}( |-|\.)?\d{2})))"  # before update


fast_try_words = ['Contact', 'Kontakt', 'お問い合わせ', 'About', 'Staff', 'Team', 'İletişim', 'Iletisim', 'iletişim', 'Bize Ulaşın', 'Hakkımızda',
                  'Contacto', 'Contatto', 'Contato', '联系', '連絡', '연락처', 'Liên hệ', 'اتصال', 'Контакт', 'Über uns', 'À propos', 'Acerca de',
                  'Informazioni', 'Sobre', '关于', '会社概要', 'について', '소개', 'Về chúng tôi', 'حول', 'О нас', 'Sobre nosotros', 'Chi siamo',
                  'Sobre nós', '关于我们', '私たちについて', '우리 소개', 'Về chúng tôi', 'عنا', 'Kontaktieren', 'Contactez', 'Contáctenos', 'Contattaci',
                  'Fale conosco', 'コミュニケーション', 'お問い合わせ', '문의하기', 'Свяжитесь с нами', 'Kommunikation', 'Communication', 'Comunicación', 'Comunicazione',
                  'Comunicação', '沟通', '커뮤니케이션', 'Giao tiếp', 'Связь']  # Word list for looking pages include these words


# Put delay
def delay(a: int, b: int):
    """Put delay"""
    time.sleep(random.uniform(a, b))


# Function for extracting emails and phones on a html
def extract_regex(content):
    """Function for extracting emails and phones on a html"""
    emails = re.findall(email_regex, content)
    phones = re.findall(phone_regex, content)
    return list(set(emails)), list(set(phones))


# Function to remove duplicates and process phone numbers to more readable shape before save
def process_phone_parts(phone_numbers):
    """Function to remove duplicates and process phone numbers to more readable shape before save"""
    separated_parts = []

    # Step 1: Separate parts if there's a space in the string
    for part in phone_numbers:
        if ' ' in part:  # If there's a space, split it into multiple parts
            separated_parts.extend(part.split())  # Split by space
        elif '-' in part:  # If there's a hyphen, split it into multiple parts
            separated_parts.extend(part.split('-'))  # Split by hyphen
        elif part:  # Keep the non-empty part as is
            separated_parts.append(part)

    # Step 2: Remove duplicates
    cleaned_parts = []
    for part in separated_parts:
        if part and part not in cleaned_parts:
            cleaned_parts.append(part)

    # Return as a tuple
    return tuple(cleaned_parts)


# Apply stealth-like techniques to the Playwright page.
def set_stealth(page):
    """Apply stealth-like techniques to the Playwright page."""
    # Set `navigator.webdriver` to `undefined`
    page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    # Fake `navigator.languages`
    page.add_init_script("""
        Object.defineProperty(navigator, 'languages', {
            get: () => ['en-US', 'en']
        });
    """)

    # Fake `navigator.plugins`
    page.add_init_script("""
        Object.defineProperty(navigator, 'plugins', {
            get: () => [1, 2, 3]
        });
    """)

    # Fake `navigator.hardwareConcurrency` to mimic real devices
    page.add_init_script("""
        Object.defineProperty(navigator, 'hardwareConcurrency', {
            get: () => 4
        });
    """)

    # Fake screen dimensions if needed
    page.add_init_script("""
        Object.defineProperty(window.screen, 'width', {get: () => 1920});
        Object.defineProperty(window.screen, 'height', {get: () => 1080});
    """)


# Function that does the search and extract
def search(search_string):
    """Function that does the search and extract"""
    with sync_playwright() as p:  # Setting browser and its settings
        browser = p.firefox.launch(headless=True)
        u_agent = random.choice(USER_AGENTS)
        print(u_agent)  # Check which User-Agent selected
        context = browser.new_context(
            user_agent=u_agent,
            viewport={"width": 1920, "height": 1080},
            locale="en-US",
            timezone_id="America/New_York"
        )
        page = context.new_page()
        set_stealth(page)

        try:
            page.goto("https://www.google.com", timeout=15000)

            # Click the button if there is one called below
            button_texts = ["Accept", "Confirm", "Ok", "I Agree", "Kabul et", "Kabul", "Tamam", "Onayla"]
            for text in button_texts:
                button = page.locator('button')
                if button.is_visible():
                    button_text = button.text_content().lower()
                    if text.lower() in button_text:
                        button.click()
                        break

            # Enter search query
            page.fill('textarea[name="q"]', search_string)
            time.sleep(random.uniform(0.5, 1))
            page.press('textarea[name="q"]', "Enter")
            page.wait_for_selector(".LC20lb", timeout=15000)

            emails = []
            phones = []

            excluded_websites = ['linkedin.com', 'youtube.com', 'facebook.com', 'yelp.com']  # Exclude these websites
            excluded_emails = ["example@email.com", "your@email.com", "email@example.com", "admin@example.com", "test@test.com", "noreply@example.com", "noreply@domain.com", "user@domain.com", "contact@domain.com", "name@domain.com"]  # Exclude these emails
            result_selector = 'a h3'  # Select this element.
            search_url = page.url

            # Select at least 7 maximum 10 unique domains. It can be reduced to 5 for faster results
            previous_domains = []
            k = 0
            for i in range(min(10, 7+k)):
                result_link = page.locator(result_selector).nth(i)
                if result_link.is_visible():
                    result_url = result_link.evaluate("element => element.closest('a').href")
                    parsed_url = urlparse(result_url)
                    domain = parsed_url.netloc
                    if previous_domains and domain in previous_domains:
                        k += 1
                        continue
                    if any(excluded in domain for excluded in excluded_websites):
                        k += 1
                        continue
                    previous_domains.append(domain)
                    print(f"Result for '{search_string}': {result_url}")
                    delay(1, 2)
                    cmp_name = page.locator('a >> span.VuuXrf').nth(i).inner_text().strip()
                    result_link.click()  # Simulating as a person really clicking the link
                    page.wait_for_load_state("load")

                    first_content = page.content()  # First page content to use later just in case

                    # Selects 50 anchor tags, click the ones that they are in the fast_try_words list above. anchor_tag.click(timeout=10000) can be reduced
                    flag = False
                    if not emails or not phones:
                        anchor_tags = page.locator('a[href]')
                        if anchor_tags.count() > 0:
                            print(f"Tags: {anchor_tags.count()}")
                            for j in range(min(anchor_tags.count(), 50)):
                                if flag:
                                    break
                                try:
                                    anchor_tag = anchor_tags.nth(j)
                                    print(f"{j}th tag: {anchor_tag}")
                                    text = anchor_tag.inner_text().strip().lower()  # Text of tag
                                    if not text:
                                        continue

                                    for word in fast_try_words:
                                        if word.lower() in text:
                                            delay(1, 2)
                                            if anchor_tag.is_visible():
                                                # anchor_tag.scroll_into_view_if_needed()
                                                print(f"Word: {word} found. Going to: {anchor_tag}")
                                                anchor_tag.click(timeout=10000)
                                                page.wait_for_load_state("load")
                                                content = page.content()

                                                soup = BeautifulSoup(content, 'lxml')  # Get page's html

                                                # Remove unwanted tags.
                                                # There can be a lot of number that a regex can capture, but they are not actual phone numbers.
                                                for tag in soup(['img', 'svg', 'script']):
                                                    tag.decompose()
                                                clean_text = soup.get_text()
                                                print(f"CLEAN TEXT: \n\n{repr(clean_text[:250])}")

                                                # Creating new email and phone instances just in case.
                                                # If there happen to be a duplication, it will be resolved in the list(set) function
                                                new_emails, new_phones = extract_regex(clean_text)
                                                emails.extend(new_emails)
                                                phones.extend(new_phones)
                                                emails, phones = list(set(emails)), list(set(phones))
                                                if emails and phones:
                                                    if any(email not in excluded_emails for email in emails):
                                                        flag = True
                                                        break
                                                    else:
                                                        page.go_back()
                                                        page.wait_for_load_state("load")
                                                else:
                                                    page.go_back()
                                                    page.wait_for_load_state("load")
                                except Exception as e:
                                    print(f"Error clicking link: {e}")

                    # If can't any find in fast_try_words, trying first page
                    if not emails or not phones:
                        soup = BeautifulSoup(first_content, 'lxml')
                        # Remove unwanted tags
                        for tag in soup(['img', 'svg', 'script']):
                            tag.decompose()
                        clean_text = soup.get_text()

                        new_emails, new_phones = extract_regex(clean_text)
                        emails.extend(new_emails)
                        phones.extend(new_phones)
                        emails, phones = list(set(emails)), list(set(phones))

                    # If still can't find, clear email and phone and go back to search_url for the next domain
                    if not emails or not phones:
                        emails.clear()
                        phones.clear()
                        page.goto(search_url)
                        page.wait_for_selector(".LC20lb", timeout=15000)
                        continue

            # If captured both return hem if not return null
            if emails and phones:
                if any(email not in excluded_emails for email in emails):
                    for _ in range(3):  # Sometimes it takes more than 1 time to fully process
                        phones = [process_phone_parts(phone) for phone in phones]
                    print(f"Emails: {emails}")
                    print(f"Phones: {phones}")
                    return cmp_name, result_url, emails, phones
                else:
                    return "", "", [], []
            else:
                return "", "", [], []

        except Exception as e:
            print(f"Error occurred during search: {e}")
            return "", "", [], []

        finally:
            delay(2, 3)
            browser.close()


# Preparing returned data to be able to saved in xlsx file.
def prepare_data(company, website, emails, phones):
    """Preparing returned data to be able to saved in xlsx file."""

    def clean_phone_number(phones):
        # Join all parts together, and then filter out unwanted characters (keeping only digits)
        phone_str = "-".join([part for part in phones if part and any(i.isdigit() for i in part)])
        return phone_str

    emails = emails[:10]
    email_str = ",".join(emails)

    phone_numbers = [clean_phone_number(phone) for phone in phones]
    phone_numbers = phone_numbers[:10]
    phone_str = ",".join(phone_numbers)  # Join phone numbers with commas

    data = [[company, website, email_str, phone_str]]
    return data


# Write data to Excel (xlsx) file
def write_to_excel(data, file_path):
    """Write data to excel (xlsx) file"""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in data:
        sheet.append(row)

    workbook.save(file_path)


# Pulling company names as a dataframe to process it
file = ""  # Path of the Excel file that contains the company names.
df = pd.read_excel(file)
company_names = df['Company Name'].dropna()
company_names = company_names[~company_names.apply(lambda x: isinstance(x, (int, float)))].astype(str).str.strip().tolist()  # Exclude numbers (false data)


# Setting xlsx file to save scraped data
scraped_excel = ""  # Path of the Excel file to save scraped data.
workbook = load_workbook(scraped_excel)
sheet = workbook.active
scraped_companies = []
failed_companies = []

# For later use, checking if there are any scraped company, if so continue scraping after the last one
for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):  # Start from row 2 to skip the header
    scraped_companies.append(row[0])  # Get the value from the first column

split_company_names = []
for name in scraped_companies:
    split_company_names.extend([n.strip() for n in name.split(',')])
scraped_companies = split_company_names

filtered_companies = [c for c in company_names if c not in scraped_companies]  # Filtering scraped company if there are any

start_time = time.time()
for company in filtered_companies[:20]:  # The number in here is to deciding how many company will be scraped. If put no number it will try to scrape all
    query = f"{company} website"
    cmp, website, emails, phones = search(query)
    if emails and phones:
        comp = company if cmp == company else f"{cmp}, {company}"  # This is to check if the scraped data is gotten through the first actual company's domain or the alternative ones. If so it will write actual wanted company then the one where the data gotten from
        data = prepare_data(comp, website, emails, phones)
        write_to_excel(data, scraped_excel)
    else:
        # If failed to scrape any, save failed companies to another Excel file
        failed_workbook = load_workbook("")  # Path of your Excel file to save failed companies
        sheet_2 = failed_workbook.active
        if isinstance(company, str):
            company = [company]  # Convert the string into a list, or adjust according to your needs.
        sheet_2.append(company)
        failed_workbook.save("")  # Path of your Excel file to save failed companies
        failed_companies.append(company)

    delay_between = random.uniform(3, 5)
    print(f"Delaying for {delay_between:.2f} seconds...")
    time.sleep(delay_between)

end_time = time.time()
print(f"Time taken: {end_time - start_time} seconds")

if failed_companies:
    print(f"Failed to retrieve: {len(failed_companies)} company data.")
